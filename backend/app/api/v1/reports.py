"""
reports.py — Reports API
"""
import csv
import io
import logging
from datetime import date, datetime, timezone, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, Integer
from app.core.database import get_db
from app.models.order import Order, OrderStatus
from app.models.algo import Algo
from app.models.account import Account

router = APIRouter()
logger = logging.getLogger(__name__)
IST = timezone(timedelta(hours=5, minutes=30))


def get_current_fy() -> str:
    from zoneinfo import ZoneInfo
    now = datetime.now(ZoneInfo("Asia/Kolkata"))
    start = now.year if now.month >= 4 else now.year - 1
    return f"{start}-{str(start + 1)[2:]}"


def _fy_range(fy: str):
    try:
        start_year = int(fy.split("-")[0])
    except Exception:
        start_year = date.today().year
    return (
        datetime(start_year,     4,  1, 0,  0,  0, tzinfo=IST),
        datetime(start_year + 1, 3, 31, 23, 59, 59, tzinfo=IST),
    )

def _base_query(fy: str, account_id: str | None, is_practix: bool | None = None):
    """Return base filter conditions for closed orders in FY."""
    fy_start, fy_end = _fy_range(fy)
    conditions = [
        Order.status == OrderStatus.CLOSED,
        Order.pnl.isnot(None),
        Order.fill_time >= fy_start,
        Order.fill_time <= fy_end,
    ]
    if account_id:
        import uuid as _uuid
        try:
            conditions.append(Order.account_id == _uuid.UUID(account_id))
        except ValueError:
            pass
    if is_practix is not None:
        conditions.append(Order.is_practix == is_practix)
    return conditions


@router.get("/metrics")
async def algo_metrics(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
):
    # When start_date/end_date are provided, use them instead of the FY range
    if start_date or end_date:
        fy_start, fy_end = _fy_range(fy)
        try:
            range_start = datetime.strptime(start_date, "%Y-%m-%d").replace(hour=0, minute=0, second=0, tzinfo=IST) if start_date else fy_start
        except ValueError:
            range_start = fy_start
        try:
            range_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=IST) if end_date else fy_end
        except ValueError:
            range_end = fy_end
        conditions = [
            Order.status == OrderStatus.CLOSED,
            Order.pnl.isnot(None),
            Order.fill_time >= range_start,
            Order.fill_time <= range_end,
        ]
        if account_id:
            import uuid as _uuid
            try:
                conditions.append(Order.account_id == _uuid.UUID(account_id))
            except ValueError:
                pass
        if is_practix is not None:
            conditions.append(Order.is_practix == is_practix)
    else:
        conditions = _base_query(fy, account_id, is_practix)
    result = await db.execute(
        select(Order, Algo)
        .join(Algo, Order.algo_id == Algo.id, isouter=True)
        .where(*conditions)
        .order_by(Order.fill_time)
    )
    rows = result.all()

    by_algo: dict = {}
    for order, algo in rows:
        aid = str(order.algo_id)
        if aid not in by_algo:
            by_algo[aid] = {"name": algo.name if algo else aid, "pnls": []}
        by_algo[aid]["pnls"].append(order.pnl or 0.0)

    metrics = []
    for aid, data in by_algo.items():
        pnls = data["pnls"]
        wins   = [p for p in pnls if p > 0]
        losses = [p for p in pnls if p <= 0]
        total  = round(sum(pnls), 2)
        metrics.append({
            "algo_id":    aid,
            "name":       data["name"],
            "trades":     len(pnls),
            "total_pnl":  total,
            "wins":       len(wins),
            "losses":     len(losses),
            "win_pct":    round(len(wins) / len(pnls) * 100, 1) if pnls else 0,
            "loss_pct":   round(len(losses) / len(pnls) * 100, 1) if pnls else 0,
            "max_profit": round(max(pnls), 2) if pnls else 0,
            "max_loss":   round(min(pnls), 2) if pnls else 0,
        })
    metrics.sort(key=lambda x: x["total_pnl"], reverse=True)

    # Fetch fy_margin for selected accounts
    if account_id:
        import uuid as _uuid
        try:
            acc_result = await db.execute(select(Account).where(Account.id == _uuid.UUID(account_id)))
            acc = acc_result.scalar_one_or_none()
            fy_margin = acc.fy_margin if acc and acc.fy_margin else 0
        except Exception as e:
            logger.warning(f"[REPORTS] Suppressed error: {e}")
            fy_margin = 0
    else:
        acc_result = await db.execute(select(Account))
        accs = acc_result.scalars().all()
        fy_margin = sum(a.fy_margin or 0 for a in accs)

    return {"metrics": metrics, "fy": fy, "fy_margin": fy_margin}


@router.get("/calendar")
async def trade_calendar(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
):
    conditions = _base_query(fy, account_id, is_practix)
    result = await db.execute(
        select(Order).where(*conditions).order_by(Order.fill_time)
    )
    orders = result.scalars().all()

    by_date: dict = {}
    for o in orders:
        if not o.fill_time:
            continue
        d = o.fill_time.astimezone(IST).date().isoformat()
        if d not in by_date:
            by_date[d] = {"pnl": 0.0, "trades": 0}
        by_date[d]["pnl"]    += o.pnl or 0.0
        by_date[d]["trades"] += 1

    calendar = [{"date": d, "pnl": round(v["pnl"], 2), "trades": v["trades"]}
                for d, v in sorted(by_date.items())]
    return {"calendar": calendar, "fy": fy}


@router.get("/equity-curve")
async def equity_curve(
    db: AsyncSession = Depends(get_db),
    fy: str | None = Query(None),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
):
    if fy is None:
        fy = get_current_fy()
    conditions = _base_query(fy, account_id, is_practix)
    result = await db.execute(
        select(Order).where(*conditions).order_by(Order.fill_time)
    )
    orders = result.scalars().all()

    by_date: dict = {}
    for o in orders:
        if not o.fill_time:
            continue
        d = o.fill_time.astimezone(IST).date().isoformat()
        by_date[d] = by_date.get(d, 0.0) + (o.pnl or 0.0)

    cumulative = 0.0
    curve = []
    for d, pnl in sorted(by_date.items()):
        cumulative += pnl
        dt = datetime.fromisoformat(d)
        curve.append({"date": d, "month": dt.strftime("%b"),
                      "pnl": round(pnl, 2), "cumulative": round(cumulative, 2)})
    return {"data": curve, "fy": fy, "total": round(cumulative, 2)}




@router.get("/day-breakdown")
async def day_breakdown(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
):
    """Day-of-week P&L breakdown per algo — powers Risk Heatmap."""
    from sqlalchemy import extract, func as sa_func
    from app.models.algo import Algo
    conditions = _base_query(fy, account_id, is_practix)
    result = await db.execute(
        select(
            Algo.name,
            extract('dow', Order.fill_time).label('dow'),
            sa_func.sum(Order.pnl).label('pnl'),
            sa_func.count(Order.id).label('trades'),
        )
        .join(Algo, Order.algo_id == Algo.id, isouter=True)
        .where(*conditions)
        .group_by(Algo.name, extract('dow', Order.fill_time))
        .order_by(Algo.name)
    )
    rows = result.all()
    DOW = {0: 'SUN', 1: 'MON', 2: 'TUE', 3: 'WED', 4: 'THU', 5: 'FRI', 6: 'SAT'}
    data: dict = {}
    for name, dow, pnl, trades in rows:
        if name not in data:
            data[name] = {}
        data[name][DOW.get(int(dow), str(dow))] = {"pnl": float(pnl or 0), "trades": int(trades)}
    return {"breakdown": data, "fy": fy}


@router.get("/errors")
async def error_analytics(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
):
    """Historical error/failure analytics — reads from event_log table.

    Returns row-level events with date, algo_name, error_type (prefix), message,
    plus by_day and by_type aggregations.
    """
    from app.models.event_log import EventLog
    from sqlalchemy import or_, cast, Date as SADate

    ERROR_PREFIXES = (
        "[ERROR]", "[MARGIN_ERROR]", "[TOKEN_ERROR]",
        "[RETRY_FAILED]", "[ENTRY_MISSED]", "[FEED_ERROR]", "[FEED]",
    )

    fy_start, fy_end = _fy_range(fy)

    conditions = [
        EventLog.ts >= fy_start,
        EventLog.ts <= fy_end,
        or_(*[EventLog.msg.like(f"{prefix}%") for prefix in ERROR_PREFIXES]),
    ]
    if account_id:
        conditions.append(EventLog.account_id == account_id)

    res = await db.execute(
        select(EventLog)
        .where(*conditions)
        .order_by(EventLog.ts.desc())
    )
    raw = res.scalars().all()

    def _extract_prefix(msg: str) -> str:
        for p in ERROR_PREFIXES:
            if msg.startswith(p):
                return p
        return "[ERROR]"

    rows = []
    by_day: dict = {}
    by_type: dict = {}

    for r in raw:
        date_str = r.ts.date().isoformat() if r.ts else None
        prefix   = _extract_prefix(r.msg or "")
        rows.append({
            "date":       date_str,
            "algo_name":  r.algo_name or "unknown",
            "error_type": prefix,
            "message":    r.msg or "",
        })
        if date_str:
            by_day[date_str] = by_day.get(date_str, 0) + 1
        by_type[prefix] = by_type.get(prefix, 0) + 1

    return {
        "rows":         rows,
        "by_day":       [{"date": d, "count": c} for d, c in sorted(by_day.items(), reverse=True)],
        "by_type":      [{"error_type": t, "count": c} for t, c in sorted(by_type.items(), key=lambda x: -x[1])],
        "total_errors": len(rows),
        "fy":           fy,
    }


@router.get("/slippage")
async def slippage_analytics(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
):
    """Historical slippage analytics — fill price vs reference price."""
    from app.models.algo import Algo
    conditions = _base_query(fy, account_id, is_practix)
    conditions.append(Order.fill_price.isnot(None))
    conditions.append(Order.entry_reference.isnot(None))

    result = await db.execute(
        select(Order, Algo.name.label('algo_name'))
        .join(Algo, Order.algo_id == Algo.id, isouter=True)
        .where(*conditions)
        .order_by(Order.fill_time.desc())
    )
    rows = result.all()

    per_algo: dict = {}
    all_slippages = []
    for r in rows:
        o = r.Order
        try:
            ref = float(o.entry_reference)
            fill = float(o.fill_price)
            slip = (fill - ref) if o.direction == 'buy' else (ref - fill)
            slip_inr = slip * (o.quantity or 1)
        except (TypeError, ValueError):
            continue

        all_slippages.append(slip)
        name = r.algo_name
        if name not in per_algo:
            per_algo[name] = {"slippages": [], "total_inr": 0.0}
        per_algo[name]["slippages"].append(slip)
        per_algo[name]["total_inr"] += slip_inr

    per_algo_out = [
        {
            "algo": name,
            "orders": len(v["slippages"]),
            "avg_slip_pts": round(sum(v["slippages"]) / len(v["slippages"]), 2),
            "total_slip_inr": round(v["total_inr"], 2),
            "best": round(min(v["slippages"]), 2),
            "worst": round(max(v["slippages"]), 2),
        }
        for name, v in per_algo.items()
    ]
    per_algo_out.sort(key=lambda x: abs(x["avg_slip_pts"]), reverse=True)

    return {
        "per_algo": per_algo_out,
        "avg_slippage_pts": round(sum(all_slippages) / len(all_slippages), 2) if all_slippages else 0,
        "total_orders_with_ref": len(all_slippages),
        "fy": fy,
    }


@router.get("/health-scores")
async def health_scores(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
):
    """Algo health scores 0-100 derived from metrics — no new DB queries beyond metrics."""
    from app.models.algo import Algo
    conditions = _base_query(fy, account_id, is_practix)

    result = await db.execute(
        select(
            Algo.name,
            func.count(Order.id).label('trades'),
            func.sum(Order.pnl).label('total_pnl'),
            func.sum(func.cast(Order.pnl > 0, Integer)).label('wins'),
        )
        .join(Algo, Order.algo_id == Algo.id, isouter=True)
        .where(*conditions)
        .group_by(Algo.name)
    )
    rows = result.all()

    scores = []
    for r in rows:
        name, trades, total_pnl, wins = r[0], int(r[1] or 0), float(r[2] or 0), int(r[3] or 0)
        losses = trades - wins
        win_pct = (wins / trades * 100) if trades > 0 else 0.0

        wr_pts   = win_pct * 0.4                             # 0-40
        pnl_pts  = min(30.0, total_pnl / 500) if total_pnl > 0 else 0  # 0-30
        con_pts  = 20 if trades >= 5 else (10 if trades >= 2 else 0)    # 0-20
        loss_pts = 10 if wins >= losses else 5                          # 5 or 10

        score = round(wr_pts + pnl_pts + con_pts + loss_pts, 1)
        grade = 'A' if score >= 80 else 'B' if score >= 60 else 'C' if score >= 40 else 'D'

        scores.append({
            "algo_name": name,
            "score": score,
            "grade": grade,
            "trades": trades,
            "win_pct": round(win_pct, 1),
            "total_pnl": round(total_pnl, 2),
            "wins": wins,
            "losses": losses,
        })

    scores.sort(key=lambda x: x["score"], reverse=True)
    avg_score = round(sum(s["score"] for s in scores) / len(scores), 1) if scores else 0
    return {"scores": scores, "avg_score": avg_score, "fy": fy}


@router.get("/all-orders")
async def all_orders(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    account_id: str | None = Query(None),
    is_practix: bool | None = Query(None),
    status: str | None = Query(None),
    limit: int = Query(500),
):
    """All historical orders without date filter — powers Analytics page."""
    from app.models.algo import Algo
    conditions = _base_query(fy, account_id, is_practix)
    if status:
        from app.models.order import OrderStatus
        try:
            conditions.append(Order.status == OrderStatus(status))
        except ValueError:
            pass

    result = await db.execute(
        select(Order, Algo.name.label('algo_name'))
        .join(Algo, Order.algo_id == Algo.id, isouter=True)
        .where(*conditions)
        .order_by(Order.fill_time.desc())
        .limit(limit)
    )
    rows = result.all()
    orders = [
        {
            "id": str(r.Order.id),
            "algo_name": r.algo_name,
            "algo_id": str(r.Order.algo_id),
            "symbol": r.Order.symbol,
            "direction": r.Order.direction,
            "status": r.Order.status.value if r.Order.status else None,
            "fill_price": float(r.Order.fill_price) if r.Order.fill_price else None,
            "entry_reference": r.Order.entry_reference,
            "exit_price": float(r.Order.exit_price) if r.Order.exit_price else None,
            "pnl": float(r.Order.pnl) if r.Order.pnl else None,
            "error_message": r.Order.error_message,
            "fill_time": r.Order.fill_time.isoformat() if r.Order.fill_time else None,
            "exit_time": r.Order.exit_time.isoformat() if r.Order.exit_time else None,
            "is_practix": r.Order.is_practix,
            "lots": r.Order.lots,
            "quantity": r.Order.quantity,
        }
        for r in rows
    ]
    return {"orders": orders, "total": len(orders), "fy": fy}

@router.get("/time-heatmap")
async def time_heatmap(
    db:         AsyncSession = Depends(get_db),
    fy:         str          = Query("2025-26"),
    account_id: str | None   = Query(None),
    is_practix: bool | None  = Query(None),
):
    """
    Group closed orders by IST entry hour (9–14).
    Returns trade count, total P&L, and win rate per hour slot.
    Used by Analytics → Best Time to Trade section.
    """
    conditions = _base_query(fy, account_id, is_practix)
    result = await db.execute(
        select(Order).where(*conditions)
    )
    orders = result.scalars().all()

    # Bucket orders by IST entry hour
    from collections import defaultdict
    buckets: dict = defaultdict(lambda: {"trades": 0, "wins": 0, "total_pnl": 0.0})
    for o in orders:
        if not o.fill_time:
            continue
        ist_hour = o.fill_time.astimezone(IST).hour
        if ist_hour < 9 or ist_hour > 14:
            continue
        b = buckets[ist_hour]
        b["trades"]    += 1
        b["total_pnl"] += float(o.pnl or 0)
        if (o.pnl or 0) > 0:
            b["wins"] += 1

    HOUR_LABELS = {9: "9–10 AM", 10: "10–11 AM", 11: "11 AM–12 PM",
                   12: "12–1 PM", 13: "1–2 PM",  14: "2–3 PM"}
    slots = []
    for hour in range(9, 15):
        b = buckets[hour]
        trades = b["trades"]
        slots.append({
            "hour":      hour,
            "label":     HOUR_LABELS[hour],
            "trades":    trades,
            "total_pnl": round(b["total_pnl"], 2),
            "win_rate":  round(b["wins"] / trades * 100, 1) if trades > 0 else 0.0,
        })

    return {"slots": slots, "fy": fy}


@router.get("/latency")
async def latency_analytics(
    db:         AsyncSession = Depends(get_db),
    fy:         str          = Query("2025-26"),
    account_id: str | None   = Query(None),
    is_practix: bool | None  = Query(None),
):
    """Order latency analytics — avg/p50/p95/max + breakdown by broker and algo."""
    from app.models.algo import Algo
    from app.models.account import Account
    fy_start, fy_end = _fy_range(fy)
    conditions = [
        Order.latency_ms.isnot(None),
        Order.placed_at >= fy_start,
        Order.placed_at <= fy_end,
    ]
    if account_id:
        import uuid as _uuid
        try:
            conditions.append(Order.account_id == _uuid.UUID(account_id))
        except ValueError:
            pass
    if is_practix is not None:
        conditions.append(Order.is_practix == is_practix)

    result = await db.execute(
        select(Order, Algo.name.label("algo_name"), Account.nickname.label("broker"))
        .join(Algo, Order.algo_id == Algo.id, isouter=True)
        .join(Account, Order.account_id == Account.id, isouter=True)
        .where(*conditions)
    )
    rows = result.all()

    if not rows:
        return {
            "avg_latency_ms": 0, "p50_latency_ms": 0, "p95_latency_ms": 0,
            "max_latency_ms": 0, "total_orders": 0,
            "by_broker": [], "by_algo": [], "fy": fy,
        }

    all_ms = sorted(int(r.Order.latency_ms) for r in rows)
    total  = len(all_ms)

    def percentile(data: list, pct: float) -> float:
        idx = max(0, int(len(data) * pct / 100) - 1)
        return float(data[idx])

    # By broker
    broker_map: dict = {}
    for r in rows:
        b = r.broker or "unknown"
        broker_map.setdefault(b, []).append(int(r.Order.latency_ms))
    by_broker = sorted([
        {"broker": b, "avg_ms": round(sum(v) / len(v), 1), "count": len(v)}
        for b, v in broker_map.items()
    ], key=lambda x: x["avg_ms"])

    # By algo
    algo_map: dict = {}
    for r in rows:
        name = r.algo_name or str(r.Order.algo_id)
        algo_map.setdefault(name, []).append(int(r.Order.latency_ms))
    by_algo = sorted([
        {
            "algo_name": name,
            "avg_ms": round(sum(v) / len(v), 1),
            "count": len(v),
            "total_orders": len(v),
        }
        for name, v in algo_map.items()
    ], key=lambda x: x["avg_ms"])

    return {
        "avg_latency_ms": round(sum(all_ms) / total, 1),
        "p50_latency_ms": percentile(all_ms, 50),
        "p95_latency_ms": percentile(all_ms, 95),
        "max_latency_ms": float(all_ms[-1]),
        "total_orders": total,
        "by_broker": by_broker,
        "by_algo":   by_algo,
        "fy":        fy,
    }


@router.get("/download")
async def download_trades(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    format: str = Query("csv"),
    account_id: str | None = Query(None),
):
    conditions = _base_query(fy, account_id)
    # download uses created_at not fill_time range — override
    fy_start, fy_end = _fy_range(fy)
    dl_conditions = [
        Order.created_at >= fy_start,
        Order.created_at <= fy_end,
        Order.status.in_([OrderStatus.CLOSED, OrderStatus.ERROR]),
    ]
    if account_id:
        import uuid as _uuid
        try:
            dl_conditions.append(Order.account_id == _uuid.UUID(account_id))
        except ValueError:
            pass
    result = await db.execute(select(Order).where(*dl_conditions).order_by(Order.created_at))
    orders = result.scalars().all()

    headers = ["Date","Symbol","Side","Qty","Entry Price","Exit Price",
               "P&L","Status","Exit Reason","Broker Order ID"]
    rows = []
    for o in orders:
        rows.append([
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
            o.symbol or "", o.direction or "", o.lots or 0,
            o.fill_price or "", o.exit_price or "",
            round(o.pnl, 2) if o.pnl is not None else "",
            o.status.value if o.status else "",
            o.exit_reason.value if o.exit_reason else "",
            o.broker_order_id or "",
        ])
    filename_base = f"STAAX_trades_FY{fy}"
    if format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"FY {fy}"
            header_fill = PatternFill("solid", fgColor="1a1a2e")
            header_font = Font(bold=True, color="00B0F0")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            return StreamingResponse(buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'})
        except ImportError:
            pass
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers); writer.writerows(rows); buf.seek(0)
    return StreamingResponse(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'})


@router.get("/strategy-breakdown")
async def strategy_breakdown(
    db: AsyncSession = Depends(get_db),
    fy: str = Query("2025-26"),
    is_practix: bool | None = Query(None),
):
    """
    Group closed trades by strategy type.
    Algos: grouped by strategy_mode (intraday/stbt/btst/positional) from Algo table.
    Bots:  grouped by indicator field (dtr/channel) from Bot table.
    Returns combined [{strategy_type, trades, total_pnl, avg_pnl, win_rate}].
    """
    groups: dict = {}

    # ── Algo orders ───────────────────────────────────────────────────────────
    conditions = _base_query(fy, None, is_practix)
    stmt = (
        select(Order, Algo.strategy_mode)
        .outerjoin(Algo, Order.algo_id == Algo.id)
        .where(*conditions)
    )
    result = await db.execute(stmt)
    for order, strategy_mode in result.all():
        stype = (strategy_mode or getattr(order, "entry_type", None) or "unknown").lower()
        if stype not in groups:
            groups[stype] = {"trades": 0, "total_pnl": 0.0, "wins": 0}
        groups[stype]["trades"] += 1
        groups[stype]["total_pnl"] += float(order.pnl or 0)
        if (order.pnl or 0) > 0:
            groups[stype]["wins"] += 1

    # ── Bot orders ────────────────────────────────────────────────────────────
    try:
        from app.models.bot import BotOrder, BotOrderStatus, Bot
        fy_start, fy_end = _fy_range(fy)
        bot_stmt = (
            select(BotOrder, Bot.indicator)
            .outerjoin(Bot, BotOrder.bot_id == Bot.id)
            .where(
                BotOrder.status == BotOrderStatus.CLOSED,
                BotOrder.pnl.isnot(None),
                BotOrder.entry_time >= fy_start,
                BotOrder.entry_time <= fy_end,
            )
        )
        bot_result = await db.execute(bot_stmt)
        for bot_order, indicator in bot_result.all():
            stype = (indicator or "unknown").lower()
            if stype not in groups:
                groups[stype] = {"trades": 0, "total_pnl": 0.0, "wins": 0}
            groups[stype]["trades"] += 1
            groups[stype]["total_pnl"] += float(bot_order.pnl or 0)
            if (bot_order.pnl or 0) > 0:
                groups[stype]["wins"] += 1
    except Exception as e:
        logger.warning(f"[REPORTS] Suppressed error: {e}")

    breakdown = [
        {
            "strategy_type": k,
            "trades": v["trades"],
            "total_pnl": round(v["total_pnl"], 2),
            "avg_pnl": round(v["total_pnl"] / v["trades"], 2) if v["trades"] > 0 else 0,
            "win_rate": round(v["wins"] / v["trades"] * 100, 1) if v["trades"] > 0 else 0,
        }
        for k, v in sorted(groups.items(), key=lambda x: -x[1]["total_pnl"])
    ]
    return {"breakdown": breakdown}
